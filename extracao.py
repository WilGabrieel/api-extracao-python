import requests
import json
import os
from datetime import datetime
import re
from openpyxl.styles import Alignment
from openpyxl import load_workbook, Workbook

# ==== Parte 1: Coletar dados da API ====
def fetch_session_data(session_id, base_url="URL-API"):
    url = f"{base_url}{session_id}"
    headers = {
        'accept': 'application/json',
        'x-api-key': 'SUA-CHAVE-API'
    }
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Erro na sessão {session_id}: {e}")
        return None

def organize_sessions_data(session_ids):
    all_sessions = {}
    for session_id in session_ids:
        data = fetch_session_data(session_id)
        if data:
            all_sessions[str(session_id)] = data
    return all_sessions

# ==== Parte 2: Ordenar por timestamp ====
def sort_messages_by_timestamp(data):
    sorted_data = {}
    for session_id, messages in data.items():
        if isinstance(messages, list):
            sorted_messages = sorted(messages, key=lambda x: x.get('timestamp', ''))
            sorted_data[session_id] = sorted_messages
        else:
            sorted_data[session_id] = messages
    return sorted_data

# ==== Parte 3: Reorganizar user/agent/function ====
def extract_text_from_content(content):
    if isinstance(content, str):
        if content.startswith('{') and content.endswith('}'):
            try:
                parsed = json.loads(content)
                if isinstance(parsed, dict) and 'params' in parsed:
                    return parsed.get('params', {}).get('motivo', content)
            except json.JSONDecodeError:
                pass
        text_match = re.search(r'<text>(.*?)</text>', content, re.DOTALL)
        if text_match:
            return text_match.group(1).strip()
        return re.sub(r'<.*?>', '', content).strip()
    return str(content)

def reorganize_conversations(data):
    reorganized_data = {}
    for session_id, messages in data.items():
        if isinstance(messages, list):
            conversation = []
            current_entry = {}
            last_sender = None
            for message in messages:
                sender = message.get('sender')
                if not sender:
                    continue
                content = extract_text_from_content(message.get('content', ''))
                if sender != last_sender and last_sender is not None:
                    conversation.append(current_entry)
                    current_entry = {}
                current_entry[sender] = content
                last_sender = sender
            if current_entry:
                conversation.append(current_entry)
            reorganized_data[session_id] = conversation
        else:
            reorganized_data[session_id] = messages
    return reorganized_data

# ==== Parte 4: Exportar para Excel com colunas customizadas ====
def json_to_excel(data, excel_file):
    wb = Workbook()
    ws = wb.active

    # Cabeçalhos
    ws.cell(row=1, column=1, value="Sessão")
    ws.cell(row=1, column=7, value="Mensagem")

    row_num = 2
    for session_id, messages in data.items():
        first = True
        for message in messages:
            for sender, content in message.items():
                sessao = session_id if first else ""
                texto = f"{sender.capitalize()}: {content}"
                ws.cell(row=row_num, column=1, value=sessao)  # Coluna A
                ws.cell(row=row_num, column=7, value=texto)   # Coluna G
                row_num += 1
                first = False

    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['G'].width = 100

    # Alinha o conteúdo
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=7):
        
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(excel_file)
    print(f"✅ Excel gerado: {excel_file}")

# ==== Util: pegar timestamp mais antigo ====
def get_first_timestamp(data):
    timestamps = []
    for messages in data.values():
        if isinstance(messages, list):
            for msg in messages:
                ts = msg.get('timestamp')
                if ts:
                    try:
                        timestamps.append(datetime.fromisoformat(ts.replace('Z', '+00:00')))
                    except Exception:
                        pass
    if timestamps:
        return min(timestamps).strftime("%Y%m%d_%H%M")
    return datetime.now().strftime("%Y%m%d_%H%M")

# ==== MAIN ====
def main():
    print("Digite os IDs das sessões separados por vírgula ou espaço:")
    input_ids = input().replace(',', ' ').split()
    try:
        session_ids = [int(id_str.strip()) for id_str in input_ids if id_str.strip()]
        print("🔄 Buscando mensagens...")
        raw_data = organize_sessions_data(session_ids)

        print("📥 Ordenando mensagens...")
        sorted_data = sort_messages_by_timestamp(raw_data)

        print("🧹 Reorganizando estrutura...")
        organized_data = reorganize_conversations(sorted_data)

        output_dir = "CAMINHO-OUTPUT"
        os.makedirs(output_dir, exist_ok=True)
        timestamp_str = get_first_timestamp(sorted_data)
        filename = os.path.join(output_dir, f"mensagens_{timestamp_str}.xlsx")

        print("📤 Salvando em Excel...")
        json_to_excel(organized_data, filename)

        print("🚀 Processo finalizado!")
    except ValueError:
        print("Erro: IDs inválidos")

if __name__ == "__main__":
    main()